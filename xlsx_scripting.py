import pandas as pd
import io
import msoffcrypto
import openpyxl

class WorkSheet:
    
    def __init__(self, filepath, password):
        self.filepath = filepath
        self.password = password
        self.wb = self.decrypt_workbook(self.filepath, self.password)
        self.main_ws = self.wb ['Instructions']
        self.list_of_worksheets = self.wb.sheetnames
        self.list_of_iterable_worksheets = self.remove_from_list(self.list_of_worksheets ,'Instructions')
        self.submitter = self.SubmitterID(self.main_ws)
        self.df = self.merge_dfs(self.list_of_iterable_worksheets, self.submitter, self.wb)

    def remove_from_list(self, list, element):
        list.remove(element)
        return list

    # IN CASE WHEN THE WORKBOOK NEEDS DECRIPTION
    def decrypt_workbook(self, filepath, password):
        decrypted_workbook = io.BytesIO()
        with open(filename, 'rb') as file:
            office_file = msoffcrypto.OfficeFile(file)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted_workbook)
        wb = openpyxl.load_workbook(filename=decrypted_workbook, read_only=True, data_only=True)
        return wb

    # EASY METHOD FOR GETTING INFORMATION FROM WORKBOOKS
    def SubmitterID(self, ws):
        sub_id = ws.cell(row=1,column=2).value
        return sub_id

    # GET DATA FROM WS RANGE AND CONVERT IT INTO PANDAS DATAFRAME
    def ws_range_to_df(self, ws , range_s = 'A1', range_e = 'H4'):
        data_rows = []
        for row in ws[range_s : range_e]:
            data_cols = []
            for cell in row:
                data_cols.append(cell.value)
            data_rows.append(data_cols)

        df = pd.DataFrame(data_rows)
        column_names = df.iloc[0]
        #print(column_names)
        df = df[1:]
        df.columns = column_names
        #print(df)
        return df

    # IN CASE YOU WANT TO LOAD THE DATA TO DATABASE, MELT THE DATAFRAME TO EASIER FORMAT
    def melt_df_to_db(self, df_raw, wsName, active = 1):
        df = pd.melt(df_raw, id_vars= ['Name/Weekday'], value_vars = [1,2,3,4,5,6,7], var_name = 'day', value_name = 'work_hours')
        df['ws'] = wsName
        #print(df)
        return df

    # MERGING DATAFRAMES CREATED FROM WORKSHEETS
    def merge_dfs(self, list_of_sheets, company_id, wb):
        n = 0
        for worksheet in list_of_sheets:
            ws = wb[worksheet]
            df_raw = self.ws_range_to_df(ws)
            n +=1
            if n == 1:
                df = self.melt_df_to_db(df_raw = df_raw, wsName = worksheet)
            else :
                df1 = self.melt_df_to_db(df_raw = df_raw, wsName = worksheet)
                df = pd.concat([df, df1], ignore_index=True)
            del df_raw
        return df

    # NEXT STEPS WOULD PROBABLY BE PUTTING THE DATAFRAME INTO DATABASE AND STORE IT THERE

filename = 'testFile.xlsx'
password = 'testTest'

a = WorkSheet(filename, password)
print(a.df)
